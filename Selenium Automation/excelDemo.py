#here we will learn how to update excel 

import openpyxl

#load_workbook - method to load the sheet , arg - put path
#went to sheet
book = openpyxl.load_workbook("E:\\DemoExcel.xlsx") #bool -obj

sheet = book.active #control of sheet obj

# cell = sheet.cell(row=1, column=2)

# print(cell.value) #O/P :- firstname 

#if we want to write something
# sheet.cell(row=2, column=2).value = 'Shruti' #writing value in excel

#trying to print what we wrote 
# print(sheet.cell(row=2, column=2).value) #Shruti

# print(f'Max row {sheet.max_row}, max column {sheet.max_column}') #Max row 2, max column 4
# print(f'Min row {sheet.min_row}, min Column {sheet.min_column}') #Min row 1, min Column 1

# print(sheet['A5'].value) #with index

#iteration using for loop
#All first colum value will update
'''
for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i, column=1).value)
'''
#O/P :-
'''
Testcase5
Testcase1
Testcase2
Testcase3
Testcase4
Testcase5
testcae6
Testcase7
'''
'''
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value) #print all the things in Excel
'''
    
#with condition 
'''
for i in range(1, sheet.max_row+1): #to get rows
    if sheet.cell(row=i, column=1).value == 'Testcase2':
        for j in range(1, sheet.max_column+1): #to get columns
            print(sheet.cell(row=i, column=j).value)
'''
#O/P :-
'''
Testcase2
a
b
c
'''
Dict = {} #emty dictionary
for i in range(1, sheet.max_row): #to get rows (iterate over rows)
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, sheet.max_column+1): #to get coulmn
            #Dict["lastname"] = "shetty"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)