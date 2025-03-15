#  Upload and Download Folder using Selenium Python...
# Generating smart XPATH to Query Table data Dynamically
# Build excel utils to update the excel file and upload it back to web portal

from selenium import webdriver
from selenium.webdriver.common.by import By

#explicitly wait
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions
import openpyxl

file_path = r"C:\Users\SHRUTI KARMAKAR\Downloads\download.xlsx"
fruit_name = input('Enter a fruit name :')
newValue = int(input('Enter a value :'))

driver = webdriver.Chrome()
driver.maximize_window()
print(driver.title)

driver.implicitly_wait(5) #implicitly wait.

driver.get('https://rahulshettyacademy.com/upload-download-test/index.html')

#click on download ...
driver.find_element(By.ID, 'downloadButton').click()


def update_excel_data(file_path, searchTerm, columnName, newValue):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {} #Empty Dictonary

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == columnName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = newValue
    book.save(file_path)

#edit the excel with updated value
update_excel_data(file_path, fruit_name,"price", newValue)


#Upload file 
file_input_button = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input_button.send_keys(file_path) #to send the file path to automate the work.
#Here the Selenium only interacts with web page not the upload folder so we are using send_keys()

#Explicitwait is required for file upload..
#waiting Till  the toast is updated and printing toast message
wait = WebDriverWait(driver, 5)

#this is locator of the Alert message of Updated folder
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")

#until() is used with wait..
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text) #* it will locok for these locator, just like page obj


#Part - 2 , Fetch the price of Apple and chenge with updated value same/not

#Fetching Price Column Dynamically
#get _attribute() will fetch value of attribute
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
print(f'Extracted price value : {priceColumn}')

#if we want to put the fruit name dynamically
#we can use f string or +, "" operator to 
# actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id ='cell-"+priceColumn+"-undefined']").text
actual_price = driver.find_element(By.XPATH, f"//div[text() = '{fruit_name}']/parent::div/parent::div/div[@id = 'cell-{priceColumn}-undefined']").text
print(f"Price of {fruit_name}: {actual_price}")
# print(f'{actual_price} = {newValue}')

driver.close() #closing the driver...

print('Updated Excel Data Successfully')

#O/P :- 1
'''
The chromedriver version (133.0.6943.141) detected in PATH at C:\WebDrivers\chromedriver.exe might not be compatible with the detected chrome version (134.0.6998.36); currently, chromedriver 134.0.6998.88 is 
recommended for chrome 134.*, so it is advised to delete the driver in PATH and retry

DevTools listening on ws://127.0.0.1:53487/devtools/browser/b7301b58-b146-4e48-b52a-7bb18d549d76        

Updated Excel Data Successfully.
---------Work Done---------
'''
'''
DevTools listening on ws://127.0.0.1:56757/devtools/browser/33c1a5a4-1a2f-4afa-ab11-1f794705b6f4

Updated Excel Data Successfully.
Extracted price value : 4
Price of Apple: 947
---------Work Done---------

'''
